from openpyxl import load_workbook
from openpyxl import Workbook
import datetime

def combine():
	dest = 'courses.xlsx'
	wb = load_workbook(filename=dest)
	stu = wb.get_sheet_by_name('students')
	time = wb.get_sheet_by_name('time')
	combine = wb.create_sheet(title='combine')
	print(wb)
	print(stu)
	print(time)

	wb.save(dest)
def split():
	pass



if __name__ == '__main__':
	combine()
	split()